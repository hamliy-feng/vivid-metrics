from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore", message="Workbook contains no default style*")

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - setup guard
    raise SystemExit("openpyxl is required to read Xiaohongshu .xlsx exports") from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT.parent.parent
OUTPUT_DIR = PROJECT_ROOT / "src" / "lib" / "data"

FIELD_VIDEO_DESC = "\u89c6\u9891\u63cf\u8ff0"
FIELD_VIDEO_ID = "\u89c6\u9891ID"
FIELD_PUBLISH_AT = "\u53d1\u5e03\u65f6\u95f4"
FIELD_COMPLETION = "\u5b8c\u64ad\u7387"
FIELD_AVG_DURATION = "\u5e73\u5747\u64ad\u653e\u65f6\u957f"
FIELD_PLAYS = "\u64ad\u653e\u91cf"
FIELD_LIKES = "\u559c\u6b22"
FIELD_COMMENTS = "\u8bc4\u8bba\u91cf"
FIELD_SHARES = "\u5206\u4eab\u91cf"
FIELD_FOLLOWERS = "\u5173\u6ce8\u91cf"
WECHAT_TRAFFIC_FIELDS = ("plays", "likes", "comments", "shares", "followers")
WECHAT_SNAPSHOT_ONLY_MODES = {"latest_file", "cheat_parallel"}

FIELD_NOTE_TITLE = "\u7b14\u8bb0\u6807\u9898"
FIELD_NOTE_TIME = "\u9996\u6b21\u53d1\u5e03\u65f6\u95f4"
FIELD_NOTE_GENRE = "\u4f53\u88c1"
FIELD_EXPOSURE = "\u66dd\u5149"
FIELD_VIEWS = "\u89c2\u770b\u91cf"
FIELD_CTR = "\u5c01\u9762\u70b9\u51fb\u7387"
FIELD_NOTE_LIKES = "\u70b9\u8d5e"
FIELD_NOTE_COMMENTS = "\u8bc4\u8bba"
FIELD_SAVES = "\u6536\u85cf"
FIELD_NOTE_FOLLOWERS = "\u6da8\u7c89"
FIELD_NOTE_SHARES = "\u5206\u4eab"
FIELD_NOTE_AVG_DURATION = "\u4eba\u5747\u89c2\u770b\u65f6\u957f"


@dataclass(frozen=True)
class SourceConfig:
    folder: str
    account: str
    patterns: tuple[str, ...]


@dataclass(frozen=True)
class FileSourceConfig:
    path: Path
    account: str
    platform_id: str
    mode: str = "latest_file"


WECHAT_SOURCES = (
    SourceConfig("vx-fans-01", "\u6b64\u5c71\u4e2d", ("*.csv",)),
    SourceConfig("vx-fans-02", "\u6ecb\u5143\u5802\u4e13\u6ce8", ("*.csv",)),
    SourceConfig("vx-new-02", "\u8001\u9ec4", ("*\u52a8\u6001\u6570\u636e\u660e\u7ec6*.csv",)),
    SourceConfig("\u4e03\u54e5\u6ecb\u8865", "\u6ecb\u5143\u5802\u6ecb\u8865", ("*.csv",)),
    SourceConfig("\u9ec4\u5e08\u5085", "\u83cc\u8bed", ("*.csv",)),
)

XHS_SOURCES = (
    SourceConfig("xhs-new-01", "\u9ec4\u5e08\u5085", ("*.xlsx",)),
    SourceConfig("xhs-new-02", "18", ("*.xlsx",)),
    SourceConfig("xhs-new-03", "\u83c7\u83c7", ("*.xlsx",)),
    SourceConfig("xhs-new-04", "\u4e8c\u80ce\u5988", ("*.xlsx",)),
)

DOWNLOADS_ROOT = Path.home() / "Downloads"
DL_ROOT = Path("D:/dl")

WECHAT_FILE_SOURCES = (
    FileSourceConfig(
        DOWNLOADS_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6 (1).csv",
        "\u6ecb\u5143\u5802\u4e13\u6ce8",
        "download-vx-fans-02",
    ),
    FileSourceConfig(
        DOWNLOADS_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6 (3).csv",
        "\u6ecb\u5143\u5802\u4e13\u6ce8",
        "download-vx-fans-02-20260611",
    ),
    FileSourceConfig(
        DL_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6.csv",
        "cheat-\u89c6\u9891\u53f7-\u672a\u5339\u914d01",
        "cheat-wechat-unknown-01",
        "cheat_parallel",
    ),
    FileSourceConfig(
        DL_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6 (1).csv",
        "cheat-\u89c6\u9891\u53f7-\u672a\u5339\u914d01",
        "cheat-wechat-unknown-01-20260611",
        "cheat_parallel",
    ),
    FileSourceConfig(
        DOWNLOADS_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6-2.csv",
        "\u83cc\u8bed",
        "download-vx-junyu",
    ),
    FileSourceConfig(
        DOWNLOADS_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6 (1)-2.csv",
        "\u6b64\u5c71\u4e2d",
        "download-vx-cishanzhong",
    ),
    FileSourceConfig(
        DATA_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6(1).csv",
        "\u83cc\u8bed",
        "data-vx-junyu-duplicate-candidate",
    ),
    FileSourceConfig(
        DATA_ROOT / "\u89c6\u9891\u53f7\u52a8\u6001\u6570\u636e\u660e\u7ec6 (5).csv",
        "\u6ecb\u5143\u5802\u6ecb\u8865",
        "data-vx-ziyuantang-zibu",
    ),
)

XHS_FILE_SOURCES = (
    FileSourceConfig(DATA_ROOT / "\u9ec4\u5e08\u5085.xlsx", "\u9ec4\u5e08\u5085", "data-xhs-huangshifu"),
    FileSourceConfig(DATA_ROOT / "18.xlsx", "18", "data-xhs-18"),
    FileSourceConfig(DATA_ROOT / "\u83c7\u83c7.xlsx", "\u83c7\u83c7", "data-xhs-gugu"),
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def to_number(value: Any, *, default: float = 0) -> float:
    text = clean_text(value)
    if not text or text == "-":
        return default
    text = text.replace(",", "").replace("\uff0c", "")
    text = re.sub(r"[\u79d2\s]", "", text)
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return default


def to_int(value: Any) -> int:
    return int(round(to_number(value, default=0)))


def parse_percent(value: Any) -> float | None:
    text = clean_text(value)
    if not text or text == "-":
        return None
    return round(to_number(text), 2)


def parse_ctr(value: Any) -> float:
    text = clean_text(value)
    if not text or text == "-":
        return 0
    if text.endswith("%"):
        return round(to_number(text) / 100, 4)
    return round(to_number(text), 4)


def parse_date(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode CSV: {path}")


def collect_files(root: Path, source: SourceConfig) -> list[Path]:
    folder = root / source.folder
    files: list[Path] = []
    for pattern in source.patterns:
        files.extend(folder.glob(pattern))
    return sorted({p for p in files if p.is_file()}, key=lambda p: (p.stat().st_mtime, str(p)))


def file_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_snapshot_at(path: Path) -> datetime:
    return datetime.fromtimestamp(path.stat().st_mtime)


def add_wechat_file(
    items: dict[tuple[str, str, str], dict[str, Any]],
    path: Path,
    account: str,
    imported: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    *,
    platform_id: str,
    mode: str,
    seen_file_hashes: dict[str, str] | None = None,
    snapshots: list[dict[str, Any]] | None = None,
) -> None:
    if not path.exists():
        skipped.append({"file": str(path), "reason": "missing_source_file", "platformId": platform_id, "mode": mode})
        return

    if seen_file_hashes is not None:
        digest = file_digest(path)
        if digest in seen_file_hashes:
            skipped.append({
                "file": str(path),
                "reason": "duplicate_source_file",
                "duplicateOf": seen_file_hashes[digest],
                "platformId": platform_id,
                "mode": mode,
            })
            return
        seen_file_hashes[digest] = str(path)

    rows = read_csv_rows(path)
    snapshot_at = file_snapshot_at(path)
    snapshot_date = snapshot_at.strftime("%Y-%m-%d")
    count = 0
    preserved_existing = 0
    added_to_base = 0
    for row in rows:
        desc = clean_text(row.get(FIELD_VIDEO_DESC))
        publish_date = parse_date(row.get(FIELD_PUBLISH_AT))
        if not desc or not publish_date:
            continue
        video_id = clean_text(row.get(FIELD_VIDEO_ID))
        item = {
            "account": account,
            "videoId": video_id,
            "desc": desc,
            "publishDate": publish_date,
            "plays": to_int(row.get(FIELD_PLAYS)),
            "likes": to_int(row.get(FIELD_LIKES)),
            "comments": to_int(row.get(FIELD_COMMENTS)),
            "shares": to_int(row.get(FIELD_SHARES)),
            "followers": to_int(row.get(FIELD_FOLLOWERS)),
            "completion": parse_percent(row.get(FIELD_COMPLETION)),
            "avgDuration": round(to_number(row.get(FIELD_AVG_DURATION)), 2),
            "snapshotDate": snapshot_date,
        }
        key = (account, video_id or desc, publish_date)
        if mode in WECHAT_SNAPSHOT_ONLY_MODES and key in items:
            preserved_existing += 1
        else:
            items[key] = item
            added_to_base += 1
        if snapshots is not None:
            snapshots.append({
                **item,
                "snapshotAt": snapshot_at.isoformat(timespec="seconds"),
                "sourcePlatformId": platform_id,
            })
        count += 1
    imported.append({
        "file": str(path),
        "account": account,
        "rows": count,
        "platformId": platform_id,
        "mode": mode,
        "baseRowsAdded": added_to_base,
        "baseRowsPreserved": preserved_existing,
    })


def build_wechat_traffic_rows(snapshots: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for snapshot in snapshots:
        key = (snapshot["account"], snapshot.get("videoId") or snapshot["desc"])
        grouped.setdefault(key, []).append(snapshot)

    traffic_rows: list[dict[str, Any]] = []
    for group in grouped.values():
        previous: dict[str, Any] | None = None
        for snapshot in sorted(group, key=lambda r: (r["snapshotAt"], r["snapshotDate"])):
            if previous is None:
                if snapshot["snapshotDate"] != snapshot["publishDate"]:
                    previous = snapshot
                    continue
                deltas = {field: snapshot[field] for field in WECHAT_TRAFFIC_FIELDS}
            else:
                deltas = {
                    field: max(0, snapshot[field] - previous.get(field, 0))
                    for field in WECHAT_TRAFFIC_FIELDS
                }

            if any(deltas.values()):
                traffic_rows.append({
                    "account": snapshot["account"],
                    "videoId": snapshot.get("videoId", ""),
                    "desc": snapshot["desc"],
                    "publishDate": snapshot["publishDate"],
                    "metricDate": snapshot["snapshotDate"],
                    "plays": deltas["plays"],
                    "likes": deltas["likes"],
                    "comments": deltas["comments"],
                    "shares": deltas["shares"],
                    "followers": deltas["followers"],
                    "completion": snapshot["completion"],
                    "avgDuration": snapshot["avgDuration"],
                    "cumulativePlays": snapshot["plays"],
                    "cumulativeLikes": snapshot["likes"],
                    "cumulativeComments": snapshot["comments"],
                    "cumulativeShares": snapshot["shares"],
                    "cumulativeFollowers": snapshot["followers"],
                })
            previous = snapshot

    return sorted(traffic_rows, key=lambda r: (r["metricDate"], r["plays"]), reverse=True)


def sync_wechat(wechat_root: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    items: dict[tuple[str, str, str], dict[str, Any]] = {}
    imported: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    seen_file_hashes: dict[str, str] = {}
    snapshots: list[dict[str, Any]] = []

    for source in WECHAT_SOURCES:
        files = collect_files(wechat_root, source)
        for path in files:
            add_wechat_file(
                items, path, source.account, imported, skipped,
                platform_id=source.folder, mode="indexed_history",
                seen_file_hashes=seen_file_hashes,
                snapshots=snapshots,
            )

    for source in WECHAT_FILE_SOURCES:
        add_wechat_file(
            items, source.path, source.account, imported, skipped,
            platform_id=source.platform_id, mode=source.mode,
            seen_file_hashes=seen_file_hashes,
            snapshots=snapshots,
        )

    configured_dirs = {source.folder for source in WECHAT_SOURCES}
    for path in sorted(wechat_root.rglob("*.csv")):
        relative = path.relative_to(wechat_root)
        if relative.parts and relative.parts[0] in configured_dirs:
            continue
        rows = read_csv_rows(path)
        desc_rows = sum(1 for row in rows if clean_text(row.get(FIELD_VIDEO_DESC)))
        skipped.append({
            "file": str(path),
            "reason": "unmapped_wechat_csv",
            "rowsWithVideoDesc": desc_rows,
        })

    data = sorted(items.values(), key=lambda r: (r["publishDate"], r["plays"]), reverse=True)
    traffic_data = build_wechat_traffic_rows(snapshots)
    return data, traffic_data, {"imported": imported, "skipped": skipped}


def iter_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows[:10]):
        values = [clean_text(cell) for cell in row]
        if FIELD_NOTE_TITLE in values and FIELD_NOTE_TIME in values:
            header_index = index
            break
    if header_index is None:
        return []

    headers = [clean_text(cell) for cell in rows[header_index]]
    parsed: list[dict[str, Any]] = []
    for row in rows[header_index + 1:]:
        record = dict(zip(headers, row))
        if clean_text(record.get(FIELD_NOTE_TITLE)):
            parsed.append(record)
    return parsed


def add_xhs_file(
    items: dict[tuple[str, str, str], dict[str, Any]],
    path: Path,
    account: str,
    imported: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    *,
    platform_id: str,
    mode: str,
    seen_file_hashes: dict[str, str] | None = None,
) -> None:
    if not path.exists():
        skipped.append({"file": str(path), "reason": "missing_source_file", "platformId": platform_id, "mode": mode})
        return

    if seen_file_hashes is not None:
        digest = file_digest(path)
        if digest in seen_file_hashes:
            skipped.append({
                "file": str(path),
                "reason": "duplicate_source_file",
                "duplicateOf": seen_file_hashes[digest],
                "platformId": platform_id,
                "mode": mode,
            })
            return
        seen_file_hashes[digest] = str(path)

    rows = iter_xlsx_rows(path)
    count = 0
    for row in rows:
        title = clean_text(row.get(FIELD_NOTE_TITLE))
        publish_date = parse_date(row.get(FIELD_NOTE_TIME))
        if not title or not publish_date:
            continue
        item = {
            "account": account,
            "title": title,
            "publishDate": publish_date,
            "exposure": to_int(row.get(FIELD_EXPOSURE)),
            "views": to_int(row.get(FIELD_VIEWS)),
            "ctr": parse_ctr(row.get(FIELD_CTR)),
            "likes": to_int(row.get(FIELD_NOTE_LIKES)),
            "comments": to_int(row.get(FIELD_NOTE_COMMENTS)),
            "saves": to_int(row.get(FIELD_SAVES)),
            "followers": to_int(row.get(FIELD_NOTE_FOLLOWERS)),
            "shares": to_int(row.get(FIELD_NOTE_SHARES)),
            "avgDuration": round(to_number(row.get(FIELD_NOTE_AVG_DURATION)), 2),
            "genre": clean_text(row.get(FIELD_NOTE_GENRE)) or "\u56fe\u6587",
        }
        key = (account, title, publish_date)
        items[key] = item
        count += 1
    imported.append({"file": str(path), "account": account, "rows": count, "platformId": platform_id, "mode": mode})


def sync_xhs(xhs_root: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    items: dict[tuple[str, str, str], dict[str, Any]] = {}
    imported: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    seen_file_hashes: dict[str, str] = {}

    for source in XHS_SOURCES:
        files = collect_files(xhs_root, source)
        for path in files:
            add_xhs_file(
                items, path, source.account, imported, skipped,
                platform_id=source.folder, mode="indexed_history",
                seen_file_hashes=seen_file_hashes,
            )

    for source in XHS_FILE_SOURCES:
        add_xhs_file(
            items, source.path, source.account, imported, skipped,
            platform_id=source.platform_id, mode=source.mode,
            seen_file_hashes=seen_file_hashes,
        )

    configured_dirs = {source.folder for source in XHS_SOURCES}
    for path in sorted(xhs_root.rglob("*.xlsx")):
        relative = path.relative_to(xhs_root)
        if relative.parts and relative.parts[0] in configured_dirs:
            continue
        skipped.append({"file": str(path), "reason": "unmapped_xhs_xlsx"})

    data = sorted(items.values(), key=lambda r: (r["publishDate"], r["views"]), reverse=True)
    return data, {"imported": imported, "skipped": skipped}


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync WeChat Video and Xiaohongshu exports into dashboard JSON.")
    parser.add_argument("--wechat-root", type=Path, default=DATA_ROOT / "wechat-video")
    parser.add_argument("--xhs-root", type=Path, default=DATA_ROOT / "xiaohongshu")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()

    if not args.wechat_root.exists():
        raise SystemExit(f"WeChat source root not found: {args.wechat_root}")
    if not args.xhs_root.exists():
        raise SystemExit(f"Xiaohongshu source root not found: {args.xhs_root}")

    wechat_data, wechat_traffic_data, wechat_report = sync_wechat(args.wechat_root)
    xhs_data, xhs_report = sync_xhs(args.xhs_root)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_json(args.output_dir / "real-wechat-data.json", wechat_data)
    write_json(args.output_dir / "wechat-traffic-data.json", wechat_traffic_data)
    write_json(args.output_dir / "xhs-real.json", xhs_data)

    report = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "wechatRows": len(wechat_data),
        "wechatTrafficRows": len(wechat_traffic_data),
        "xhsRows": len(xhs_data),
        "wechatAccounts": sorted({row["account"] for row in wechat_data}),
        "wechatCheatAccounts": sorted({row["account"] for row in wechat_data if row["account"].startswith("cheat-")}),
        "xhsAccounts": sorted({row["account"] for row in xhs_data}),
        "wechat": wechat_report,
        "xiaohongshu": xhs_report,
    }
    write_json(args.output_dir / "source-sync-report.json", report)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
